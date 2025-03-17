import os
import hashlib
import json
import re
import subprocess
import tempfile
import logging

import torch
import numpy as np

# Milvus / PyMilvus
from pymilvus import (
    connections,
    Collection,
    CollectionSchema,
    FieldSchema,
    DataType,
    Index,
    IndexType,
    utility
)

# Neo4j
from neo4j import GraphDatabase

# UI / Progress
from halo import Halo
from tqdm import tqdm

# PDF reading
from PyPDF2 import PdfReader

# doc reading
try:
    import docx
except ImportError:
    docx = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None

try:
    import whisper
except ImportError:
    whisper = None

# Language models for embeddings (GPU enabled if available)
from sentence_transformers import SentenceTransformer

# Multilingual summarization pipeline (GPU enabled if available)
try:
    from transformers import pipeline
    device_for_hf = 0 if torch.cuda.is_available() else -1
    # Example multilingual summarizer (mT5 trained on XLSum)
    summarizer = pipeline("summarization", model="csebuetnlp/mT5_multilingual_XLSum", device=device_for_hf)
except Exception as e:
    summarizer = None
    print("Could not load multilingual summarization pipeline. Install 'transformers' or check GPU setup. Error:", e)

# spaCy for NER and optional chunking (English only by default)
import spacy
try:
    if torch.cuda.is_available():
        spacy.require_gpu()
        print("spaCy GPU enabled.")
except Exception as e:
    print("spaCy GPU enable failed:", e)

try:
    nlp_spacy_en = spacy.load("en_core_web_sm")
    # nlp_spacy_en.to("cuda")  # optional: If you want to push spaCy model to GPU
except OSError:
    nlp_spacy_en = None
    print("spaCy 'en_core_web_sm' model not found. Install with: python -m spacy download en_core_web_sm")

# LangChain-based text splitter
from langchain.text_splitter import RecursiveCharacterTextSplitter

# Simple language detection for deciding whether to do English NER
try:
    from langdetect import detect
    can_detect_language = True
except ImportError:
    can_detect_language = False
    print("langdetect not installed. pip install langdetect if you want dynamic language detection.")

###############################################################################
# Setup Logging
###############################################################################

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

###############################################################################
# Global variables
###############################################################################

driver = None

# Two embedding models for multi-vector approach
semantic_embedding_model = None
lexical_embedding_model = None

# Two Milvus collections
semantic_collection = None
lexical_collection = None

# Dimensions for each model
semantic_dim = None
lexical_dim = None

###############################################################################
# Functions to Initialize DB / Models
###############################################################################

def init_neo4j(
    neo4j_uri="bolt://localhost:7687",
    neo4j_user="neo4j",
    neo4j_password="testtest"
):
    """
    Initializes the Neo4j driver.
    """
    global driver
    driver = GraphDatabase.driver(neo4j_uri, auth=(neo4j_user, neo4j_password))
    logger.info(f"Connected to Neo4j at {neo4j_uri}")


def init_semantic_embedding_model(model_name="sentence-transformers/paraphrase-multilingual-mpnet-base-v2"):
    """
    Loads a "semantic" model from SentenceTransformers on GPU if available.
    Using a multilingual model for broader coverage.
    """
    global semantic_embedding_model, semantic_dim
    logger.info(f"Loading SEMANTIC embedding model '{model_name}'...")

    device_str = "cuda" if torch.cuda.is_available() else "cpu"
    with Halo(text=f"Loading semantic embedding model on {device_str}...", spinner="dots") as spinner:
        semantic_embedding_model = SentenceTransformer(model_name, device=device_str)
        spinner.succeed(f"Semantic model '{model_name}' loaded successfully (device: {device_str}).")

    semantic_dim = semantic_embedding_model.get_sentence_embedding_dimension()
    logger.info(f"Detected semantic model dimension: {semantic_dim}")


def init_lexical_embedding_model(model_name="sentence-transformers/all-mpnet-base-v2"):
    """
    Loads a second model for "lexical" or specialized embeddings on GPU if available.
    """
    global lexical_embedding_model, lexical_dim
    logger.info(f"Loading LEXICAL embedding model '{model_name}'...")

    device_str = "cuda" if torch.cuda.is_available() else "cpu"
    with Halo(text=f"Loading lexical embedding model on {device_str}...", spinner="dots") as spinner:
        lexical_embedding_model = SentenceTransformer(model_name, device=device_str)
        spinner.succeed(f"Lexical model '{model_name}' loaded successfully (device: {device_str}).")

    lexical_dim = lexical_embedding_model.get_sentence_embedding_dimension()
    logger.info(f"Detected lexical model dimension: {lexical_dim}")


def create_or_load_milvus_collection(collection_name, dim, host="localhost", port="19530"):
    """
    Creates or loads a Milvus collection of specified dimension.
    """
    connections.connect("default", host=host, port=port)
    if collection_name not in utility.list_collections():
        fields = [
            FieldSchema(
                name="document_id",
                dtype=DataType.INT64,
                is_primary=True,
                auto_id=False
            ),
            FieldSchema(
                name="embedding",
                dtype=DataType.FLOAT_VECTOR,
                dim=dim
            ),
        ]
        schema = CollectionSchema(fields, description=f"{collection_name} schema")

        with Halo(text=f"Creating Milvus collection '{collection_name}'...", spinner="dots") as spinner:
            coll = Collection(name=collection_name, schema=schema)
            spinner.succeed(f"Collection '{collection_name}' created.")

        # Example: HNSW index
        index_params = {
            "index_type": IndexType.HNSW,
            "metric_type": "COSINE",
            "params": {"M": 16, "efConstruction": 200},
        }
        with Halo(text=f"Creating index for '{collection_name}'...", spinner="dots") as spinner:
            Index(coll, "embedding", index_params)
            spinner.succeed("Index created successfully.")

        coll.load()
        return coll
    else:
        with Halo(text=f"Loading existing collection '{collection_name}'...", spinner="dots") as spinner:
            coll = Collection(collection_name)
            coll.load()
            spinner.succeed(f"Collection '{collection_name}' loaded.")
        return coll


def initialize_all(
    neo4j_uri="bolt://localhost:7687",
    neo4j_user="neo4j",
    neo4j_password="testtest",
    milvus_host="localhost",
    milvus_port="19530"
):
    """
    Sets up the embedding models, Milvus collections, and Neo4j driver.
    """
    logger.info("Initializing embedding models...")

    # Switch to a multilingual semantic model
    init_semantic_embedding_model("sentence-transformers/paraphrase-multilingual-mpnet-base-v2")
    # Possibly keep or change lexical model as well
    init_lexical_embedding_model("sentence-transformers/all-mpnet-base-v2")

    logger.info("Creating/Loading Milvus collections for multi-vector embeddings...")
    global semantic_collection, lexical_collection
    semantic_collection = create_or_load_milvus_collection(
        "document_embeddings_semantic",
        dim=semantic_dim,
        host=milvus_host,
        port=milvus_port
    )
    lexical_collection = create_or_load_milvus_collection(
        "document_embeddings_lexical",
        dim=lexical_dim,
        host=milvus_host,
        port=milvus_port
    )

    logger.info("Initializing Neo4j connection...")
    init_neo4j(neo4j_uri=neo4j_uri, neo4j_user=neo4j_user, neo4j_password=neo4j_password)
    logger.info("All services initialized successfully.")

###############################################################################
# Neo4j / Graph Helpers
###############################################################################

def clear_neo4j_graph(batch_size=10):
    """
    Clears the entire Neo4j graph in batches to avoid memory/time issues.
    """
    with driver.session() as session:
        with Halo(text="Clearing Neo4j graph in batches...", spinner="dots") as spinner:
            deleted = batch_size
            total_deleted = 0
            while deleted == batch_size:
                result = session.run(f"""
                MATCH (n)
                WITH n LIMIT {batch_size}
                DETACH DELETE n
                RETURN count(*) AS count
                """)
                record = result.single()
                deleted = record["count"] if record else 0
                total_deleted += deleted
                if deleted == 0:
                    break
            spinner.succeed(f"Neo4j graph cleared successfully. Total nodes deleted: {total_deleted}")


def _create_document_node_tx(tx, doc_id, content, metadata_json):
    """
    Creates a Document node.
    """
    tx.run(
        """
        CREATE (d:Document {
            doc_id: $doc_id,
            content: $content,
            metadata: $metadata_json
        })
        """,
        doc_id=doc_id,
        content=content,
        metadata_json=metadata_json
    )

def create_document_node(doc_id, content, metadata):
    """
    Insert doc node into Neo4j. Make sure all numeric data is convertible to JSON.
    """
    safe_metadata = {}
    for k, v in metadata.items():
        if isinstance(v, (np.integer, np.floating)):
            safe_metadata[k] = float(v)
        else:
            safe_metadata[k] = v

    metadata_json = json.dumps(safe_metadata, ensure_ascii=False)
    with driver.session() as session:
        session.write_transaction(_create_document_node_tx, doc_id, content, metadata_json)
    logger.debug(f"Created Neo4j node for doc_id: {doc_id}")

def _create_relationship_tx(tx, doc_id_1, doc_id_2, relationship_type, extra_data):
    tx.run(
        """
        MATCH (d1:Document {doc_id: $doc_id_1})
        WITH d1
        MATCH (d2:Document {doc_id: $doc_id_2})
        WHERE d1 <> d2
        CREATE (d1)-[:RELATED {type: $relationship_type, extra: $extra_data}]->(d2)
        """,
        doc_id_1=doc_id_1,
        doc_id_2=doc_id_2,
        relationship_type=relationship_type,
        extra_data=extra_data
    )

def create_relationship(doc_id_1, doc_id_2, relationship_type, extra_data=None):
    """
    We must cast any np.float32 to float before JSON serialization.
    """
    if extra_data and "score" in extra_data:
        extra_data["score"] = float(extra_data["score"])

    extra_data_str = json.dumps(extra_data or {})

    with driver.session() as session:
        session.write_transaction(
            _create_relationship_tx,
            doc_id_1,
            doc_id_2,
            relationship_type,
            extra_data_str
        )
    logger.debug(f"Created relationship '{relationship_type}' between {doc_id_1} and {doc_id_2}")

def _create_topic_node_tx(tx, topic_id, topic_summary):
    tx.run(
        """
        CREATE (t:Topic {
            topic_id: $topic_id,
            summary: $topic_summary
        })
        """,
        topic_id=topic_id,
        topic_summary=topic_summary
    )

def create_topic_node(topic_id, topic_summary):
    with driver.session() as session:
        session.write_transaction(_create_topic_node_tx, topic_id, topic_summary)
    logger.debug(f"Created Topic node for id: {topic_id}")

def _create_has_topic_tx(tx, doc_id, topic_id):
    tx.run(
        """
        MATCH (d:Document {doc_id: $doc_id})
        MATCH (t:Topic {topic_id: $topic_id})
        CREATE (d)-[:HAS_TOPIC]->(t)
        """,
        doc_id=doc_id,
        topic_id=topic_id
    )

def create_has_topic_relationship(doc_id, topic_id):
    with driver.session() as session:
        session.write_transaction(_create_has_topic_tx, doc_id, topic_id)
    logger.debug(f"Created HAS_TOPIC relationship from doc:{doc_id} to topic:{topic_id}")

def _create_entity_node_tx(tx, entity_id, entity_text, entity_label):
    tx.run(
        """
        MERGE (e:Entity {entity_id: $entity_id})
        ON CREATE SET e.name = $entity_text, e.label = $entity_label
        """,
        entity_id=entity_id,
        entity_text=entity_text,
        entity_label=entity_label
    )

def _create_mentions_rel_tx(tx, doc_id, entity_id):
    tx.run(
        """
        MATCH (d:Document {doc_id: $doc_id})
        MATCH (e:Entity {entity_id: $entity_id})
        MERGE (d)-[:MENTIONS]->(e)
        """,
        doc_id=doc_id,
        entity_id=entity_id
    )

def create_entity_node_and_rel(doc_id, entity_text, label):
    """
    Creates or merges an Entity node, then links Document -> Entity with MENTIONS.
    """
    entity_id = int(hashlib.sha256(entity_text.encode("utf-8")).hexdigest(), 16) % (2 ** 63 - 1)
    with driver.session() as session:
        session.write_transaction(_create_entity_node_tx, entity_id, entity_text, label)
        session.write_transaction(_create_mentions_rel_tx, doc_id, entity_id)


###############################################################################
# Summaries & Entity Extraction
###############################################################################

def summarize_chunk_with_llm(text):
    """
    Summarizes text using the Hugging Face pipeline (mT5_multilingual_XLSum) on GPU if available.
    If the summarizer is unavailable, we do a fallback naive approach.
    """
    if summarizer:
        try:
            # For mT5_multilingual_XLSum, you might adapt max_length, min_length depending on doc language
            result = summarizer(text, max_length=100, min_length=30, do_sample=False)
            summary = result[0]["summary_text"]
            return summary.strip()
        except Exception as e:
            logger.warning(f"Summarization error: {e}. Using fallback naive summary.")

    # Fallback if summarizer not available
    max_len = 120
    naive_summary = text.strip().replace("\n", " ")
    if len(naive_summary) > max_len:
        naive_summary = naive_summary[:max_len] + "..."
    return naive_summary

def detect_language_of_text(text):
    """
    Simple helper using langdetect if installed. Returns 'en' or iso code, or None if detection fails.
    """
    if not can_detect_language:
        return None
    try:
        lang_code = detect(text)
        return lang_code
    except:
        return None

def extract_entities_and_relationships(text):
    """
    We do English-only spaCy-based NER if the text is detected as English.
    For other languages, we skip NER.
    If you want fully multilingual NER, load a model that supports those languages.
    """
    lang_code = detect_language_of_text(text)
    if lang_code != "en":
        # skip NER for non-English
        logger.debug(f"Skipping spaCy NER; doc language detected: {lang_code}")
        return []

    if not nlp_spacy_en:
        return []

    doc = nlp_spacy_en(text)
    entities = []
    for ent in doc.ents:
        entities.append((ent.text, ent.label_))
    return entities


###############################################################################
# Embedding & Similarity
###############################################################################

def generate_doc_id(content):
    sha256_hash = hashlib.sha256(content.encode('utf-8')).hexdigest()
    doc_id = int(sha256_hash, 16) % (2 ** 63 - 1)
    return doc_id

def store_embeddings_in_collection(doc_ids, embeddings, milvus_collection):
    embeddings = np.array(embeddings, dtype=np.float32)
    milvus_collection.insert([doc_ids, embeddings.tolist()])
    milvus_collection.flush()

def store_multi_vector_embeddings(doc_id, chunk_text):
    """
    Generates both semantic and lexical embeddings, stores them in separate collections.
    GPU is used if available (set when loading the models).
    """
    # SEMANTIC
    sem_emb = semantic_embedding_model.encode([chunk_text], show_progress_bar=False)[0]
    sem_emb = sem_emb.astype(np.float32)
    store_embeddings_in_collection([doc_id], [sem_emb], semantic_collection)

    # LEXICAL
    lex_emb = lexical_embedding_model.encode([chunk_text], show_progress_bar=False)[0]
    lex_emb = lex_emb.astype(np.float32)
    store_embeddings_in_collection([doc_id], [lex_emb], lexical_collection)

def cosine_similarity(vec1, vec2):
    return np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2))

def compute_batch_similarities(doc_ids, embeddings, relationship_type="SIMILAR_TO", threshold=0.7):
    """
    For each embedding, compare with the previous ones in the batch,
    link them if similarity > threshold.
    Make sure to cast similarity to float.
    """
    for i in range(len(embeddings)):
        for j in range(i):
            sim_val = cosine_similarity(embeddings[i], embeddings[j])
            if sim_val > threshold:
                create_relationship(
                    doc_ids[i],
                    doc_ids[j],
                    relationship_type,
                    extra_data={"score": float(sim_val)}
                )

###############################################################################
# Chunking Approaches
###############################################################################

def chunk_text_with_spacy(doc_text, min_chunk_size=500, max_chunk_size=2000):
    """
    spaCy-based chunking by sentences (English only).
    If doc is non-English, we won't do a language-based approach; you could skip or use a different model.
    """
    # Just for demonstration, we do not do language detection here.
    # If you want, detect language, pick the right spaCy model or fallback, etc.

    if not nlp_spacy_en:
        logger.warning("English spaCy not available, using fallback chunking.")
        return [doc_text]

    doc = nlp_spacy_en(doc_text)
    chunks = []
    current_chunk = []
    current_length = 0

    for sent in doc.sents:
        sentence_text = sent.text.strip()
        if not sentence_text:
            continue

        if (current_length + len(sentence_text)) > max_chunk_size and current_length >= min_chunk_size:
            chunks.append(" ".join(current_chunk))
            current_chunk = [sentence_text]
            current_length = len(sentence_text)
        else:
            current_chunk.append(sentence_text)
            current_length += len(sentence_text)

    # Add the remainder
    if current_chunk:
        chunks.append(" ".join(current_chunk))

    logger.debug(f"SPACY chunked text into {len(chunks)} chunks.")
    return chunks

def chunk_text_with_langchain(doc_text, chunk_size=1000, chunk_overlap=200):
    splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=chunk_overlap)
    chunks = splitter.split_text(doc_text)
    logger.debug(f"LangChain chunked text into {len(chunks)} chunks.")
    return chunks

###############################################################################
# File Reading
###############################################################################

def transcribe_video(file_path, whisper_model_size="base"):
    if not whisper:
        logger.warning("Whisper not installed. Please `pip install openai-whisper` if needed.")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        audio_path = os.path.join(tmpdir, "temp_audio.wav")
        ffmpeg_cmd = [
            "ffmpeg",
            "-i", file_path,
            "-ar", "16000",
            "-ac", "1",
            "-f", "wav",
            "-y",
            audio_path
        ]

        try:
            subprocess.run(ffmpeg_cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            logger.debug(f"Extracted audio to {audio_path}")
        except subprocess.CalledProcessError as e:
            logger.error(f"Error extracting audio: {e}")
            return None

        model = whisper.load_model(whisper_model_size)
        logger.info(f"Transcribing video {file_path} with Whisper model '{whisper_model_size}'...")
        result = model.transcribe(audio_path)
        return result["text"].strip() if "text" in result else None

def read_file_content(file_path):
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".md":
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()

        elif ext == ".pdf":
            try:
                reader = PdfReader(file_path)
                text = "\n".join([page.extract_text() or "" for page in reader.pages])
                return text
            except Exception as e:
                logger.error(f"Error reading PDF {file_path}: {e}")
                return None

        elif ext == ".docx" and docx:
            try:
                doc = docx.Document(file_path)
                return "\n".join(p.text for p in doc.paragraphs)
            except Exception as e:
                logger.error(f"Error reading DOCX {file_path}: {e}")
                return None

        elif ext in [".xlsx", ".xls"] and load_workbook:
            try:
                wb = load_workbook(filename=file_path)
                text_content = []
                for sheet in wb:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " ".join(str(cell) for cell in row if cell is not None)
                        text_content.append(row_text)
                return "\n".join(text_content)
            except Exception as e:
                logger.error(f"Error reading XLSX {file_path}: {e}")
                return None

        elif ext in [".png", ".jpg", ".jpeg", ".tif", ".tiff"] and Image and pytesseract:
            try:
                image = Image.open(file_path)
                text = pytesseract.image_to_string(image)
                return text.strip() if text else None
            except Exception as e:
                logger.error(f"Error reading image {file_path}: {e}")
                return None

        elif ext in [".mp4", ".mov", ".avi", ".mkv"]:
            return transcribe_video(file_path)

        else:
            # fallback to plain text read
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()

    except Exception as e:
        logger.error(f"An error occurred reading {file_path}: {e}")
        return None

def extract_metadata(file_path, content):
    metadata = {
        "filename": os.path.basename(file_path),
        "size": os.path.getsize(file_path),
        "word_count": len(content.split()) if content else 0,
    }
    return metadata

###############################################################################
# Processing & Ingestion
###############################################################################

def process_chunk(doc_text, file_metadata):
    """
    Main logic for a single chunk:
    1. Generate doc_id
    2. Summarize chunk -> create Topic node (multilingual summarizer)
    3. Create Document node in Neo4j with chunk + metadata
    4. (English-only) Extract entities + create MENTIONS relationships
    5. Store multi-vector embeddings in Milvus
    """
    doc_id = generate_doc_id(doc_text)

    # Summarize chunk text -> create topic node
    topic_summary = summarize_chunk_with_llm(doc_text)
    topic_id = generate_doc_id(topic_summary)
    create_topic_node(topic_id, topic_summary)

    # Add summary to the metadata
    meta_copy = dict(file_metadata)
    meta_copy["chunk_summary"] = topic_summary

    # Create Document node
    create_document_node(doc_id, doc_text, meta_copy)

    # Link Document -> Topic
    create_has_topic_relationship(doc_id, topic_id)

    # Entity Extraction -> create MENTIONS relationships (English only here)
    entity_list = extract_entities_and_relationships(doc_text)
    for ent_text, ent_label in entity_list:
        create_entity_node_and_rel(doc_id, ent_text, ent_label)

    # Multi-vector embeddings (semantic + lexical)
    store_multi_vector_embeddings(doc_id, doc_text)

    return doc_id

def process_documents(directory, batch_size=50, use_spacy_chunking=False):
    """
    1. Clears Neo4j graph (fresh start).
    2. Reads files, chunking them with either spaCy or LangChain.
    3. For each chunk, processes it (document node, topic node, entity nodes, embeddings).
    4. Batches doc-doc similarity computation for the "SIMILAR_TO" edges.
    """
    clear_neo4j_graph()

    all_chunks = []
    logger.info(f"Reading documents from: {directory}")

    for filename in tqdm(os.listdir(directory), desc="Scanning files"):
        file_path = os.path.join(directory, filename)
        if not os.path.isfile(file_path):
            continue

        content = read_file_content(file_path)
        if not content:
            logger.warning(f"No content extracted: {file_path}")
            continue

        # Decide chunk approach
        if use_spacy_chunking:
            # Only really appropriate for English text with the loaded model
            chunks = chunk_text_with_spacy(content, min_chunk_size=500, max_chunk_size=2000)
        else:
            chunks = chunk_text_with_langchain(content, chunk_size=1000, chunk_overlap=200)

        file_meta = extract_metadata(file_path, content)

        for c in chunks:
            all_chunks.append((c, file_meta))

    logger.info(f"Total chunks to process: {len(all_chunks)}")

    doc_ids_batch = []
    embeddings_semantic_batch = []

    # Process each chunk
    for i, (chunk_text, file_meta) in enumerate(tqdm(all_chunks, desc="Processing chunks")):
        doc_id = process_chunk(chunk_text, file_meta)
        doc_ids_batch.append(doc_id)

        # For doc-doc similarities, we store the semantic embedding
        sem_emb = semantic_embedding_model.encode([chunk_text], show_progress_bar=False)[0]
        embeddings_semantic_batch.append(sem_emb)

        # Every 'batch_size' chunks, compute doc-doc similarity
        if (i + 1) % batch_size == 0 or (i + 1) == len(all_chunks):
            logger.info(f"Computing doc-doc similarities for batch up to index {i}...")
            compute_batch_similarities(doc_ids_batch, embeddings_semantic_batch, threshold=0.7)
            doc_ids_batch.clear()
            embeddings_semantic_batch.clear()

###############################################################################
# MAIN
###############################################################################

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Process and embed documents with multi-vector approach (GPU-enabled, multilingual).")
    parser.add_argument(
        "--directory",
        type=str,
        default="path/to/documents",
        help="Directory containing documents to process."
    )
    parser.add_argument(
        "--batch_size",
        type=int,
        default=50,
        help="Number of chunks for doc-doc similarity computations in each batch."
    )
    parser.add_argument(
        "--spacy_chunking",
        action="store_true",
        help="Use spaCy-based chunking (English only) instead of LangChain text splitter."
    )
    parser.add_argument(
        "--log_level",
        type=str,
        default="INFO",
        help="Logging level (DEBUG, INFO, WARNING, ERROR, CRITICAL)."
    )
    args = parser.parse_args()

    # Set logging level
    numeric_level = getattr(logging, args.log_level.upper(), None)
    if not isinstance(numeric_level, int):
        logger.error(f"Invalid log level: {args.log_level}")
        exit(1)
    logger.setLevel(numeric_level)

    # Validate directory
    if not os.path.isdir(args.directory):
        logger.error(f"Invalid directory: {args.directory}")
        exit(1)

    # Initialize everything
    initialize_all()

    with Halo(text="Starting ingestion...", spinner="dots") as spinner:
        try:
            process_documents(
                directory=args.directory,
                batch_size=args.batch_size,
                use_spacy_chunking=args.spacy_chunking
            )
            spinner.succeed("Document processing completed.")
        except Exception as e:
            spinner.fail(f"Error in processing: {e}")
